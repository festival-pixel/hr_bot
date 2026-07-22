from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.constants import STATUS_NAMES, VACANCY_NAMES
from app.database.models import Candidate

HEADERS = [
    "№", "Заявка", "Дата", "Статус", "Вакансия", "График", "ФИО",
    "Возраст", "Телефон", "Студент", "Стаж", "Последнее место",
    "Дети", "Детей", "Возр. младшего", "Адрес", "Языки", "О себе",
    "Резюме", "Username", "Геолокация",
]

COLUMN_WIDTHS = [
    5, 12, 18, 14, 22, 14, 26, 8, 16, 9, 18, 24, 8, 7, 14,
    26, 18, 34, 9, 16, 22,
]


def build_candidates_xlsx(candidates: list[Candidate]) -> BytesIO:
    """Формирует .xlsx со всеми заявками в память."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for idx, c in enumerate(candidates, start=1):
        geo = f"{c.latitude}, {c.longitude}" if c.latitude else "—"
        ws.append([
            idx,
            c.application_number,
            c.created_at.strftime("%d.%m.%Y %H:%M"),
            STATUS_NAMES.get(c.status, c.status),
            VACANCY_NAMES["ru"].get(c.vacancy, c.vacancy),
            c.schedule,
            c.fullname,
            c.age,
            c.phone,
            "Да" if c.student else "Нет",
            c.experience,
            c.last_workplace,
            "Да" if c.has_children else "Нет",
            c.children_count if c.has_children else "—",
            c.youngest_child_age if c.has_children else "—",
            c.address,
            c.languages,
            c.motivation,
            ("фото" if c.resume_type == "photo" else "документ")
            if c.resume_file_id
            else "нет",
            f"@{c.username}" if c.username else "—",
            geo,
        ])

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
